"""
aviv drori
may-june 2021
This file uses opnenpyxl to read and write to the excel
"""

from openpyxl import load_workbook
from datetime import datetime, time, date
import config
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import get_column_letter

# EXCEL FILE CONSTANTS:
database_file = config.get_database_file_name()  # this file needs to be in same dir

# WRITE HISTORY:
SHEET_HISTORY_NAME = 'WriteHistory'
LAST_WRITE_CELL = 'G3'

# COL_DAYS
COL_FROM_DICT = {1: 'A', 2: 'E', 3: 'I', 4: 'M', 5: 'Q', 6: 'U', 7: 'Y'}
COL_TO_DICT = {1: 'B', 2: 'F', 3: 'J', 4: 'N', 5: 'R', 6: 'V', 7: 'Z'}
COL_NAME_DICT = {1: 'C', 2: 'G', 3: 'K', 4: 'O', 5: 'S', 6: 'W', 7: 'AA'}
COL_LEN_DICT = {1: 'D', 2: 'H', 3: 'L', 4: 'P', 5: 'T', 6: 'X', 7: 'AB'}
# ROWS
ROW_EVENTS_START = 7
ROW_EVENTS_END = 15
# CLEAR RANGE
CLEAR_RANGE = 'A' + str(ROW_EVENTS_START) + ':AB' + str(ROW_EVENTS_END)
# EXCEL SHEET STARTUP
DB = load_workbook(config.get_database_file_name())
SHEET_HISTORY = DB[SHEET_HISTORY_NAME]

SHEET_DICT = {1: DB['week1'], 2: DB['week2'], 3: DB['week3'],
              4: DB['week4'], 5: DB['week5'], 6: DB['week6'],
              7: DB['week7'], 8: DB['week8'], 9: DB['week9'],
              10: DB['week10'], 11: DB['week11'], 12: DB['week12'],
              13: DB['week13'], 14: DB['week14'], 15: DB['week15'],
              16: DB['week16'], 17: DB['week17'], 18: DB['week18'],
              19: DB['week19'], 20: DB['week20'], 21: DB['week21'],
              22: DB['week22'], 23: DB['week23'], 24: DB['week24'],
              25: DB['week25'], 26: DB['week26']}

ALL_WEEKS = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14,
             15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26]
weeks_written = []
# FILTER
bad_words = ["ללא לימודים", "יום הולדת"]  # default
FILTERED_SHEET_NAME = 'filtered'
SHEET_FILTERED = DB[FILTERED_SHEET_NAME]
FILTER_LAST_ROW = 'H2'
# COLORS
COLOR_TO_CODE = {'Yellow': '00FFFF00', 'Red': '00FF0000', 'Cyan': '0000FFFF',
                 'Purple': '007030A0', 'Green': '003DEB07', 'White': '00FFFFFF'}
colors_keywords = config.read_color_file()
COLOR_DAYS135 = '00D9E1F2'
COLOR_DAYS24 = '00FCE4D6'
COLUMN_DAY_DICT = {'A': 1, 'B': 1, 'C': 1, 'D': 1,
                   'E': 2, 'F': 2, 'G': 2, 'H': 2,
                   'I': 3, 'J': 3, 'K': 3, 'L': 3,
                   'M': 4, 'N': 4, 'O': 4, 'P': 4,
                   'Q': 5, 'R': 5, 'S': 5, 'T': 5,
                   'U': 6, 'V': 6, 'W': 6, 'X': 6,
                   'Y': 7, 'Z': 7, 'AA': 7, 'AB': 7}
COLOR_WKND_HOME = '0092D050'
COLOR_WKND_BASE = '00FFC000'
WEEKEND_HOME = 'יוצאים'
WEEKEND_BASE = 'סוגרים'
WKND_RANGE_TO_COLOR = 'U3:AB21'


def setup_settings(database_file_name, new_bad_words):
    global database_file
    database_file = database_file_name
    global bad_words
    bad_words = new_bad_words
    global colors_keywords
    colors_keywords = config.read_color_file()
    print(bad_words)


def clear_events(weeks):
    """
    Clears all the events written to the sheet in CLEAR_RANGE.
    :param weeks: list of week numbers to clear
    :return: None
    """
    for week_num in weeks:
        for row in SHEET_DICT[week_num][CLEAR_RANGE]:
            for cell in row:
                cell.value = None
                fill = None
                day = COLUMN_DAY_DICT[get_column_letter(cell.column)]
                if day in [1, 3, 5]:
                    fill = PatternFill(start_color=COLOR_DAYS135,
                                       end_color=COLOR_DAYS135,
                                       fill_type='solid')
                if day in [2, 4]:
                    fill = PatternFill(start_color=COLOR_DAYS24,
                                       end_color=COLOR_DAYS24,
                                       fill_type='solid')
                if fill is not None:
                    cell.fill = fill


def check_and_color(event, i):
    """
    :param event: event obj
    :param i: row of excel sheet to color
    :return:
    """
    summary = event['summary']
    for key in colors_keywords:
        keywords = colors_keywords[key]
        for word in keywords:
            if word in summary:
                color_event(event, i, COLOR_TO_CODE[key])


def color_event(event, i, color_code):
    """
    set the color of the event to color_code (RGB string)
    :param event: event object
    :param i: row
    :param color_code: RGB string like '00FFFF00'
    :return: None
    """
    dtm_start, dtm_end = event_to_datetime(event)
    sheet = choose_sheet(dtm_start)
    day = choose_day(dtm_start)
    i_ = str(i)
    fill = PatternFill(start_color=color_code,
                       end_color=color_code,
                       fill_type='solid')
    sheet[COL_NAME_DICT[day] + i_].fill = fill
    sheet[COL_TO_DICT[day] + i_].fill = fill
    sheet[COL_FROM_DICT[day] + i_].fill = fill
    sheet[COL_LEN_DICT[day] + i_].fill = fill


def weekend_check_and_color(event):
    """
    checks if stay on base/go home and sets the color accordingly
    :param event: event obj that is a full day event (filtered)
    :return: None
    """
    if WEEKEND_HOME in event['summary']:
        use = COLOR_WKND_HOME

    elif WEEKEND_BASE in event['summary']:
        use = COLOR_WKND_BASE
    else:
        return
    fill = PatternFill(start_color=use,
                       end_color=use,
                       fill_type='solid')
    sheet = choose_sheet(dtm_whole_day(event))
    for row in sheet[WKND_RANGE_TO_COLOR]:
        for cell in row:
            if not (cell.fill.bgColor.rgb == COLOR_DAYS24 or COLOR_DAYS135):
                cell.fill = fill
            else:
                if cell.value is None:
                    cell.fill = fill


def dtm_whole_day(event):
    """
    for whole day events
    :param event: full day event obj
    :return: the datetime of it (date is correct, random time)
    """
    # like 2021-05-31
    event_str = event['start'].get('dateTime', event['start'].get('date'))
    date_obj = date(int(event_str[0:4]), int(event_str[5:7]),
                    int(event_str[8:10]))
    time_obj = time(1, 1)
    return datetime.combine(date_obj, time_obj)


def should_write(event_obj):
    """
    checks if event contains a forbidden substring like "without studies",
     or whole day/zero time
    :param event_obj: event object
    :return: BOOL - True if event should be written, False otherwise
    """
    start = event_obj['start'].get('dateTime', event_obj['start'].get('date'))
    dtm_start = dtm_str_to_obj(start)
    correct_sheet = choose_sheet(dtm_start)

    if correct_sheet == 0:
        weekend_check_and_color(event_obj)
        print("    Event not written: Filtered: whole day event")
        write_to_filtered_list(event_obj, "whole day event")
        return False
    event_summary = event_obj['summary']
    for bad_word in bad_words:
        if bad_word in event_summary:
            print("    Event not written: Filtered: contains string ",
                  bad_word)
            write_to_filtered_list(event_obj, "contains string" + bad_word)
            return False
    return True


def write_to_filtered_list(event_obj, why):
    """
    writes the filtered events to the 'filtered' worksheet
    :param event_obj: the object that was filtered
    :param why: string representing reason event was filtered
    :return: None
    """
    row = int(SHEET_FILTERED[FILTER_LAST_ROW].value)
    row += 1
    row_ = str(row)
    dtm_start, dtm_end = event_to_datetime(event_obj)
    SHEET_FILTERED["A" + row_] = event_obj['summary']
    SHEET_FILTERED["B" + row_] = choose_week(dtm_start)
    SHEET_FILTERED["C" + row_] = str(dtm_start.date())
    SHEET_FILTERED["D" + row_] = str(dtm_start.time())[0:5]
    SHEET_FILTERED["E" + row_] = why
    SHEET_FILTERED[FILTER_LAST_ROW] = row


def dtm_str_to_obj(dtm_string):
    """
    :param dtm_string: string representing datetime, from google calendar
    :return: dtm: datetime object of the input
    """
    if len(dtm_string) == 10:   # like 2021-05-31
        # print("Full-day event.")
        return datetime(2020, 1, 1, 1, 1, 1)  # arbitrary time
    # time_ = time_string[11:19]
    time_obj = time(int(dtm_string[11:13]), int(dtm_string[14:16], 0))
    # print(time_obj)
    date_obj = date(int(dtm_string[0:4]), int(dtm_string[5:7]),
                    int(dtm_string[8:10]))
    # print(date_obj)
    dtm = datetime.combine(date_obj, time_obj)
    # print(dtm)
    return dtm


def choose_sheet(dtm_event):
    week_num = choose_week(dtm_event)
    if week_num == 0:
        return 0
    if week_num not in SHEET_DICT:
        return 0
    return SHEET_DICT[week_num]


def choose_week(dtm_event):
    """
    :param dtm_event: datetime object of event
    :return: int representing air force week of the date
    """
    if dtm_event == datetime(2020, 1, 1, 1, 1, 1):
        # this is the arbitrary datetime for whole day events
        return 0
    week_num = dtm_event.isocalendar()[1] + 1
    if choose_day(dtm_event) == 1:
        week_num += 1  # small error fix for sundays
    while week_num > 26:
        week_num -= 26
    if week_num == 0:
        week_num = 1
    return week_num


def choose_day(dtm_event):
    """
    returns the day of the week properly
    :param dtm_event:
    :return: 1 for sunday, 2 for monday...
    """
    day = dtm_event.date().isoweekday()
    if day == 7:
        day = 0
    day += 1
    return day


def write_event(event, i):
    """
    :param event: event obj to write to the file
    :param i: integer - the row to write to
    :return:
    """
    dtm_start, dtm_end = event_to_datetime(event)
    elapsed_time = (dtm_end - dtm_start).seconds / 3600  # hours
    sheet = choose_sheet(dtm_start)
    if sheet == 0:
        write_to_filtered_list(event, "No matching sheet for this week")
        return
    # prepare
    day = choose_day(dtm_start)
    i_ = str(i)
    week_num = choose_week(dtm_start)
    # clear
    global weeks_written
    if week_num not in weeks_written:
        if not len(weeks_written) == 0:
            clear_events([week_num])
        weeks_written.append(week_num)
    # write from, to, name, length
    sheet[COL_FROM_DICT[day] + i_] = str(dtm_start.time())[0:5]
    sheet[COL_TO_DICT[day] + i_] = str(dtm_end.time())[0:5]
    sheet[COL_NAME_DICT[day] + i_] = str(event['summary'])
    sheet[COL_LEN_DICT[day] + i_] = elapsed_time
    # output
    print("  Writing to sheet %s, wk:%d, day:%d, hr:%s, len:%d"
          % (sheet, week_num, day, str(dtm_start.time())[0:5], elapsed_time))


def event_to_datetime(event):
    """
    Parses the info, just for cleaner code. JSON style.
    :param event: event object
    :return: datetime of start, datetime of end
    """
    start = event['start'].get('dateTime', event['start'].get('date'))
    end = event['end'].get('dateTime', event['end'].get('date'))
    dtm_start = dtm_str_to_obj(start)
    dtm_end = dtm_str_to_obj(end)
    return dtm_start, dtm_end


def log(how_many_events):
    """
    logs to the WriteHistory sheet
    :return: None
    """
    row = int(SHEET_HISTORY[LAST_WRITE_CELL].value)
    row += 1
    row_ = str(row)
    SHEET_HISTORY["A" + row_] = str(datetime.now())
    SHEET_HISTORY["B" + row_] = how_many_events
    SHEET_HISTORY[LAST_WRITE_CELL] = row
    SHEET_FILTERED[FILTER_LAST_ROW] = 1  # ready for next time
    print("Successfully logged to history.")


def write_events_to_db(events):
    """
    receives list of event and manages the filtering and writing.
    :param events: list of events as received from google calendar
    :return: n/a
    """
    events_written = 0
    i = ROW_EVENTS_START
    prev_day = 0
    for event in events:
        print("Received event: ", event['summary'])
        if should_write(event):
            dtm_start, dtm_end = event_to_datetime(event)
            day = choose_day(dtm_start)
            if day != prev_day:
                i = ROW_EVENTS_START
                prev_day = day
            if i == ROW_EVENTS_END:
                print("ERROR: Overflow: Too many events to write.")
                write_to_filtered_list(event, "OVERFLOW")
            else:
                write_event(event, i)
                check_and_color(event, i)
                i += 1
                events_written += 1
    print("--------------------------------------------------------------")
    print("Done receiving all events.")
    log(len(events))
    save()
    print("Finished.\nGot ", len(events),
          " events, wrote ", events_written, " of them.")


def save():
    try:
        DB.save(database_file)
        print("Excel database saved.")
    except PermissionError:
        input("ERROR! Make sure excel database is closed, then press enter.")
        save()
    finally:
        return


def tests():
    # tests:
    save()


if __name__ == '__main__':
    tests()
